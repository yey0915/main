# 엑셀의 파일의 확장자 xls -> xlsx
# pip install openpyxl
# 터미널 설치
import openpyxl

# xlsx 읽은 인스턴스 workbook
workbook = openpyxl.load_workbook('c:/CookAnalysis/Excel/singer.xlsx')
# 엑셀의 워크시트 이름 목록
wsheetList = workbook.sheetnames

# 워크시트 목록에서 하나의 워크시트 이름 선택하는 반복문
for wsName in wsheetList :
    worksheet = workbook[wsName]
    print('** 워크시트의 이름 : %s' % (wsName) )
    # 끝자리가 미포함, (현태, 헤더 + 데이터 같이 포함. )
    for row in range(1, worksheet.max_row+1) :
        # 열도 하나 증가시켜서, 반복 실행. (끝까지 미포함. )
        for col in range(1, worksheet.max_column+1) :
            # 엑셀의 각 셀의 위치의 값을 조회
            print("%s" % (worksheet.cell(row=row, column=col).value), end='\t')
        print()
    print()