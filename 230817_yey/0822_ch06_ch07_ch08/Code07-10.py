import openpyxl
# xlsx
# 엑셀 파일 읽은 인스턴스
from copy import copy

workbook = openpyxl.load_workbook('c:/CookAnalysis/Excel/singer.xlsx')
# 엑셀의 워크시트 목록
wsheetList = workbook.sheetnames

# xlsx 파일을 읽기 위한 인스턴스
outWorkbook = openpyxl.Workbook()
# 읽은 엑셀 파일에서, 해당 워크시트의 기본 시트 제거.
outWorkbook.remove(outWorkbook['Sheet']) # 기본으로 생성된 시트를 일단 제거

# 읽은 엑셀 파일의 워크시트 이름 이용해서, 작업.
for wsName in wsheetList :
    # wsName : 읽은 엑셀 파일의 워크시트
    # worksheet : wsName의 내용
    worksheet = workbook[wsName]
    outSheet = outWorkbook.create_sheet(wsName)
    # 각 행, 열 이중 for 문 이용해서, 각 셀의 값을 채우는 내용.
    for row in range(1, worksheet.max_row+1) :
        # 새 엑셀 파일의 행의 높이 속성, 높이는 기존과 동일
        # worksheet : 읽은 엑셀 파일
        # outsheet : 새 엑셀 파일
        outSheet.row_dimensions[row].height = worksheet.row_dimensions[row].height
        # 행에 대한 열의 값 설정
        for col in range(1, worksheet.max_column+1) :
            # 새 엑셀 파일의 열이름(A~Z) 값을 지정
            # 새 액셀 파일의 열의 넓일ㄹ 동일하게
            print(f"chr(ord('A')+col-1) 의 결과 : {chr(ord('A')+col-1)}")
            outSheet.column_dimensions[chr(ord('A')+col-1)].width \
                = worksheet.column_dimensions[chr(ord('A')+col-1)].width
            # inCell : 읽은 엑셀 파일이 행, 열의 위치의 값
            inCell = worksheet.cell(row=row, column=col)
            # outCell : 새 엑셀 파일에 해당 셀의 값을 지정
            outCell = outSheet.cell(row=row, column=col, value= inCell.value)
            # 엑셀 각 셀마다, 속성, 각셀의 크기, 폰트, 서식(모양), 배경색
            if inCell.has_style:
                # 기존 엑셀의 셀의 속성 모양을, 새 엑셀의 셀 모양(서식 복사.)
                outCell.font = copy(inCell.font)
                outCell.border = copy(inCell.border)
                outCell.fill = copy(inCell.fill)
                outCell.number_format = copy(inCell.number_format)
                outCell.alignment = copy(inCell.alignment)

outWorkbook.save('c:/CookAnalysis/Excel/singer_copy_0822.xlsx')
print("Save. OK~")